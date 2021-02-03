#  ===========================
# -*- coding:utf-8 -*-
# Time :2020/12/19 9:47
# Author :A0025-江苏-小丹
# QQ:915155536
# File :data_driver.py
#  ===========================
import openpyxl
from sss.Api.base_api import WebApi
from sss.my_log.my_log import my_log

# 构建日志
logger = my_log('data_driver')
my_book = openpyxl.load_workbook('1.xlsx')
sheet1 = my_book['data_sheet']
for value in sheet1.values:
    if isinstance(value[0], str):
        pass
    elif value[0] == 1:
        logger.info(f'已打开{value[5]}浏览器')
        my_browser = WebApi(value[5])
    else:
        logger.info(f'开始执行：<<{value[2]}>> 步骤')
        # 判断参数是否为空，不为空则加入列表，生成要传入的参数
        args = []
        for i in range(3, 7):
            if value[i] == None:
                pass
            else:
                args.append(value[i])
        # 把参数列表转换成元组
        args = tuple(args)
        #断言用例，则写入断言结果
        if 'assert' in value[1]:
            if getattr(my_browser,value[1])(*args):
                sheet1.cell(value[0]+1,8).value='Pass'
            else:
                sheet1.cell(value[0]+1,8).value='failed'
            my_book.save('1.xlsx')
        #普通用例：
        else:
            getattr(my_browser, value[1])(*args)
            logger.info(f'<<{value[2]}>>步骤执行成功！')
