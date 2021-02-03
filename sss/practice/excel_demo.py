#  ===========================
# -*- coding:utf-8 -*-
# Time :2020/11/4 14:49
# Author :A0025-江苏-小丹
# QQ:915155536
# File :excel_demo.py
#  ===========================
import openpyxl

#获取工作簿
wk=openpyxl.load_workbook('1.xlsx')

#获取工作簿中的所有工作表
print(wk.sheetnames)

#获取工作簿中某一工作表
sheet1=wk['Sheet1']
print(sheet1)

#获取表中某一单元格(row：行号，column：列号，均从1开始)
cell=sheet1.cell(row=1,column=1)
print(cell)
#获取单元格的值及坐标
print(cell.value,cell.coordinate)

#创建一个工作簿
my_workbook=openpyxl.workbook.Workbook()
#创建一个工作表
my_sheet=my_workbook.create_sheet('my_sheet')
#往单元格写入数据
my_sheet.cell(1,1,'序号')
my_sheet.cell(1,2,'课程')
my_sheet.cell(2,1,1)
my_sheet.cell(2,2,'语文1111')
my_sheet.cell(2,3).value='hahahahha'
my_workbook.save('2.xlsx')
 

